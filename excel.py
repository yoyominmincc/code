import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
wb.create_sheet('mysheet')

wb['A2'].value='dddddd'
ws.title = "test"
ws['A1'].value= 'cccccc'
wb.save(r'test1.xlsx')